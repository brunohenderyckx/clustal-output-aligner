from flask import Flask, Response, render_template, request
from openpyxl import Workbook
from openpyxl.styles import Font, colors
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series


def soft_check(cell, ref_cell, matching_rules=[("D", "E", "N", "Q"),
                                               ("K", "R", "H"),
                                               ("F", "W", "Y"),
                                               ("V", "I", "L", "M"),
                                               ("S", "T")]):
    """ Checks if the cell value matches the matching list of tuples of the reference cell

    Args:
        cell (string): A row of Clustal output
        ref_cell (string): A row of Clustal output
        matching_dict (dictionary): A row of Clustal output

    Returns:
        Bool: Returns True if the values or soft matches
    """
    for group in matching_rules:
        if cell in group and ref_cell in group:
            return True
    else:
        return False


def create_matching_dict(matching_input):
    # creating the matching dictionary
    if matching_input == "":
        matching_rules = [("D", "E", "N", "Q"),
                          ("K", "R", "H"),
                          ("F", "W", "Y"),
                          ("V", "I", "L", "M"),
                          ("S", "T")]
    else:
        matching_rules = []
        for line in matching_input.replace('\n',"").split("\r"):
            temp = ()
            for letter in line:
                temp = temp + tuple(letter)
            matching_rules.append(temp)
    return matching_rules


def row_protein_length(row):
    """ Returns length of protein  string in Clustal rows

    Args:
        row (list): A row of Clustal output

    Returns:
        length: An integer for the number of protein characters in the row
    """
    length = 0
    for index, character in enumerate(row):
        if index != 0 and \
            (character.isupper() or character == "-") and \
                (" " in row[:index]):
            length += 1

    return length


def convert_raw_clustal(source_data, cut_of_number=5):
    """
    Converts a clustal output string into a list of rows we want to work with.
    Empty lines or lines with only non-alphabetical characters get removed.
    Cut of number determines how many characters the rows minimally need to have to be considered.
    """

    # takes the Clustal data and splits it into a list of rows if there's at least 5 characters
    source_data = source_data.splitlines()
    source_data = [x for x in source_data if len(x) > cut_of_number]

    # removes the list item if there's no upper case character
    rows_to_remove = []
    for x in source_data:
        if any(ele.isupper() for ele in x) == False:
            rows_to_remove.append(x)

    # need to remove the list items if they are in the index list
    source_data = [item for item in source_data if item not in rows_to_remove]

    return source_data


def protein_aligner_single(alignment_input, matching_rules):
    """
    Generates the excel file with the data and color the alignment in a single row per protein
    """
    PADDING_LEFT = 1
    PADDING_RIGHT = 1
    PADDING_BOTTOM = 1
    PADDING_TOP = 1
    EXCEL_OFFSET = 1
    stft = Font(name='Consolas', size=10.5)
    ft = Font(name='Consolas', size=10.5, color=colors.WHITE)
    HARD_COLOR = "494544"
    SOFT_COLOR = "7A7675"

    # generates the Clustal data
    source_data = convert_raw_clustal(alignment_input, 5)

    # create and activate excel worksheet
    workbook = Workbook()
    sheet = workbook.active

    # figures out what the column-indices are for the alignment comparisons
    column_range = []
    for index, character in enumerate(source_data[0]):
        if index != 0 and \
            (character.isupper() or character == "-") and \
                (" " in source_data[0][:index]):  # only upper characters or -
            column_range.append(index)

    # generate a dictionary with unique series names and empty lists to store source_data rows in
    species = {}
    for row in source_data:
        name = row[0:min(column_range)].strip()
        if name and not (name in species):
            species[name] = []

    # save the names to column A
    for i, key in enumerate(species):
        sheet.cell(row=i + PADDING_TOP + EXCEL_OFFSET, column=1).value = key
        sheet.cell(row=i + PADDING_TOP + EXCEL_OFFSET, column=1).font = stft
        species[key].append(int(i) + PADDING_TOP + EXCEL_OFFSET)

    # generate all row indices and loop over the rows
    for i in range(0, len(source_data)):

        # saves the species_name of the row we are at
        # so we can check the dictionary
        row_name = source_data[i][0:min(column_range)].strip()

        if row_name not in species:  # checks if the row name is in the dictionary or not
            continue

        excel_row = min(species[row_name])
        series_rows_processed = len(species[row_name])

        # generates a range from the minimum index in our protein column range
        # to the maximum + 1 index except if the row is smaller than that
        #
        for j in range(min(column_range),
                       min(
            max(column_range) + 1,
            row_protein_length(source_data[i]) + min(column_range))
        ):

            # save the output to excel with the standard font
            # now we don't need to increment the row
            # we instead need to add it to the row in the dictionary
            try:
                char_to_write = source_data[i][j]

                col_to_write = j + EXCEL_OFFSET + \
                    (len(column_range) * (series_rows_processed - 1)) - \
                    (min(column_range) - 1)

                sheet.cell(row=excel_row,
                           column=col_to_write).value = char_to_write
                sheet.cell(row=excel_row, column=col_to_write).font = stft
                
            except:
                pass


        species[row_name].append(int(i + EXCEL_OFFSET + PADDING_RIGHT))

    """
    Colouring based on matching for Excel Cells
    """
    # how many columns?
    last_empty_column = len(list(sheet.columns))

    
    for i in range(EXCEL_OFFSET + 1, len(species) + EXCEL_OFFSET + 1):
        row_match_counter = [0, 0, 0]
        for j in range(2, last_empty_column + EXCEL_OFFSET):
            cell = sheet.cell(row=i, column=j).value
            ref_cell = sheet.cell(row=2, column=j).value

            # compare if hard match
            if cell == ref_cell and ref_cell != "-":
                sheet.cell(row=i, column=j).fill = PatternFill(
                    fgColor=HARD_COLOR, fill_type="solid")
                sheet.cell(row=i, column=j).font = ft
                row_match_counter[0] += 1

            # compare if soft match
            elif soft_check(cell, ref_cell, matching_rules) and ref_cell != "-":
                sheet.cell(row=i, column=j).fill = PatternFill(
                    fgColor=SOFT_COLOR, fill_type="solid")
                sheet.cell(row=i, column=j).font = ft
                row_match_counter[1] += 1
            
            # all the other matches
            elif ref_cell != "-" and cell == "-":
                row_match_counter[2] += 1
            elif ref_cell == "-" and cell == "-":
                pass
            elif ref_cell != "-" and cell == "-":
                pass
            else:
                row_match_counter[2] += 1

        # save counter aggregates reference species
        for index, value in enumerate(row_match_counter):
            sheet.cell(row=i, column=last_empty_column +
                       EXCEL_OFFSET + PADDING_RIGHT + index).value = value

        sheet.cell(row=i, column=last_empty_column + EXCEL_OFFSET +
                   PADDING_RIGHT + 3).value = sum(row_match_counter)

        
        CHART_OFFSET = 5
        for index, value in enumerate(row_match_counter):
            sheet.cell(row=i, column=last_empty_column +
                       EXCEL_OFFSET + PADDING_RIGHT + index + CHART_OFFSET).value = value / sum(row_match_counter)

    # """
    # Create plot
    # """
    # values = Reference(sheet, min_col=582 + CHART_OFFSET, min_row=2, max_col= 582 + 2 + CHART_OFFSET, max_row=3)
    # cats = Reference(sheet, min_col=1, min_row=2, max_col= 1, max_row=3)
    # chart = BarChart()
    # chart.add_data(values)
    # chart.set_categories(cats)
    # chart.type = "bar"
    # chart.grouping = "percentStacked"
    # chart.title = 'Percent Stacked Chart'
    # sheet.add_chart(chart, "VJ6")

    """
    Format the rest of the excel file
    """
    # print Count header
    sheet.cell(row=1, column=1).value = "Name"

    header_counter_names = ['Identical', 'Conserved', 'Variable', 'Total']
    for i, header_name in enumerate(header_counter_names):
        sheet.cell(row=1, column=last_empty_column + EXCEL_OFFSET +
                   PADDING_RIGHT + i).value = header_name

    header_counter_names = ['% Identical', '% Conserved', '% Variable']
    for i, header_name in enumerate(header_counter_names):
        sheet.cell(row=1, column=last_empty_column + EXCEL_OFFSET +
                   PADDING_RIGHT + i + CHART_OFFSET).value = header_name

    # set the width of the columns to optimize viewing display
    i = get_column_letter(1)
    sheet.column_dimensions[i].width = 20
    sheet.column_dimensions[i].width = 20

    column = 2
    while column < (last_empty_column + len(header_counter_names) + 1):

        if column > last_empty_column + PADDING_RIGHT:
            i = get_column_letter(column)
            sheet.column_dimensions[i].width = 12
            column += 1

        else:
            i = get_column_letter(column)
            sheet.column_dimensions[i].width = 1.5
            column += 1

    sheet.sheet_view.showGridLines = False
    return workbook
